import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def export_to_excel():
    print("正在讀取合併後的數據...")
    with open("merged_l1_data.json", "r", encoding="utf-8") as f:
        data = json.load(f)
        
    print(f"總計讀取 {len(data)} 個單字。")
    
    # Prepare list for DataFrame
    rows = []
    for word, details in data.items():
        rows.append({
            "Vocabulary": word,
            "IPA": details.get("ipa", ""),
            "Part of Speech / Definition": details.get("def", ""),
            "Inflection": details.get("trans", ""),
            "Collocation": details.get("col", ""),
            "Example Sentence": details.get("ex", "")
        })
        
    df = pd.DataFrame(rows)
    
    output_path = "Level_1_Vocabulary_Enriched_Full_AZ.xlsx"
    
    # Use ExcelWriter for styling
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Vocabulary')
        
        workbook = writer.book
        worksheet = writer.sheets['Vocabulary']
        
        # Styling
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # Apply styles to header
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
        # Set column widths
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 40
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 35
        worksheet.column_dimensions['F'].width = 60
        
        # Apply borders and alignment to data cells
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.border = thin_border
                if cell.column == 1 or cell.column == 2:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment

    print(f"\nExcel 匯出完成！檔案路徑: {output_path}")

if __name__ == "__main__":
    export_to_excel()
